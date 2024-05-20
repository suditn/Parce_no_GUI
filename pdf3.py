import os
import pdfplumber
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter
import re

ser1 = []
ser2 = []

def there_is_pdf(folder_path, excel_file, excel_sheet):
    # Загружаем существующий файл Excel
    wb = load_workbook(excel_file)
    ws = wb[excel_sheet]

    # Создаем словарь для хранения путей к файлам по сериям
    series_paths = {}

    # Рекурсивно сканируем папки и обновляем данные в Excel
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith((".pdf")):
                file_path = os.path.join(root, file)
                # Вычисляем относительный путь
                relative_path = os.path.relpath(file_path, folder_path)
                ser1.append(relative_path)
                # Получаем название серии из пути к файлу
                series_name = os.path.dirname(relative_path)
                series_name = series_name.split(os.sep)[-1]  # Удаляем "Datasheet\" из имени серии
                ser2.append(series_name)
                # Добавляем путь к файлу в словарь для соответствующей серии
                if series_name not in series_paths:
                    series_paths[series_name] = []
                if relative_path not in series_paths[series_name]:
                    series_paths[series_name].append(relative_path)

# Функция для создания и записи DataFrame в Excel
def create_part_number(file_name, impedance, tolerance):
    parts = file_name.split('-')
    if len(parts) < 2:
        return None

    product_family = parts[0]
    size = parts[1]

    # Формируем IMPEDANCE из значения
    if impedance is not None:
        try:
            impedance_float = float(str(impedance).replace(',', '.'))
            if impedance_float < 1:
                impedance_value = str(impedance_float).replace('0.', 'R')
            elif impedance_float < 10:
                impedance_value = str(impedance_float).replace('.', 'R')
            else:
                impedance_value = str(int(impedance_float))
                if len(impedance_value) == 2:
                    impedance_value += '0'
                elif len(impedance_value) == 3:
                    impedance_value = impedance_value[:-1] + '1'
                elif len(impedance_value) == 4:
                    impedance_value = impedance_value[:-2] + '2'
        except ValueError:
            impedance_value = "000"
    else:
        impedance_value = "000"

    # Определяем TOLERANCE
    if tolerance == 20:
        tol = "M"
    elif tolerance == 30:
        tol = "N"
    else:
        tol = ""

    # Составляем базовый PART NUMBER
    part_number = f"{product_family}{size}ER{impedance_value}{tol}"

    # Добавляем часть после третьего тире, если она существует
    if len(parts) > 2:
        part_number += parts[2]

    return part_number

# Функция для установки ширины столбцов
def set_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

# Функция для создания и записи DataFrame в Excel
def create_excel_with_file_description(save_path, pdf_path):
    # Извлечение названия файла без расширения
    file_name = os.path.basename(pdf_path).replace(".pdf", "")

    table_data = []
    column_names = []

    # Открытие PDF-файла и извлечение таблиц
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if table and len(table) > 1 and table[0][0] == "STANDARD ELECTRICAL SPECIFICATIONS":
                    if not column_names:
                        column_names = table[1]
                    table_data.extend(table[2:])

    if not column_names:
        max_columns = max(len(row) for row in table_data)
        column_names = [f"Column {i+1}" for i in range(max_columns)]

    # Преобразование данных таблицы в числовые значения с заменой точки на запятую
    for row_index, row in enumerate(table_data):
        for col_index, value in enumerate(row):
            if value is not None:
                try:
                    if isinstance(value, str):
                        value = value.replace('.', ',')
                    table_data[row_index][col_index] = float(value.replace(',', '.'))
                except (ValueError, TypeError):
                    table_data[row_index][col_index] = value

    # Открытие PDF-файла и извлечение текста
    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += page.extract_text()

    lines = [line for line in full_text.splitlines() if line.strip()]

    # Создание DataFrame из таблиц и добавление столбцов
    df = pd.DataFrame(table_data, columns=column_names)

    df["File Name"] = file_name
    df["Description"] = "\n".join(lines[1:6]) if len(lines) > 1 else ""

    # Поиск столбца с "IMPEDAN" или "INDUCTAN" в названии
    impedance_column = None
    for col in df.columns:
        if col and ("IMPEDAN" in col.upper() or "INDUCTAN" in col.upper()):
            impedance_column = col
            break

    # Проверка наличия столбца PART NUMBER и добавление его при необходимости
    if impedance_column and "PART NUMBER" not in df.columns:
        df["PART NUMBER"] = df.apply(lambda row: create_part_number(
            file_name,
            row[impedance_column],
            20  # Замените это значение на актуальное значение tolerance, если оно имеется в таблице
        ), axis=1)

    # Перемещение столбцов "File Name", "Description" и "PART NUMBER" в начало DataFrame
    columns = ["File Name", "Description"]
    if "PART NUMBER" in df.columns:
        columns.append("PART NUMBER")
    columns += [col for col in df.columns if col not in columns]

    df = df[columns]

    # Запись DataFrame в Excel
    output_excel_path = os.path.join(save_path,"Datasheet",file_name, f"{file_name}.xlsx")
    output_csv_path = os.path.join(save_path,"Datasheet",file_name, f"{file_name}.csv")
    df.to_excel(output_excel_path, index=False)
    df.to_csv(output_csv_path, index=False)

    return output_excel_path

# Функция для обновления и форматирования файла Excel
def update_excel_with_file_paths(excel_file, excel_sheet):
    wb = load_workbook(excel_file)
    ws = wb[excel_sheet]

    first_cell_font = Font(
        name=ws.cell(row=2, column=1).font.name,
        size=ws.cell(row=2, column=1).font.sz,
        bold=ws.cell(row=2, column=1).font.b,
        italic=ws.cell(row=2, column=1).font.i,
        color=ws.cell(row=2, column=1).font.color,
        underline=ws.cell(row=2, column=1).font.underline,
        strike=ws.cell(row=2, column=1).font.strike,
    )

    merged_cell_style_name = "merged_cell_style"
    if merged_cell_style_name not in wb.named_styles:
        merged_cell_style = NamedStyle(name=merged_cell_style_name, font=first_cell_font)
        merged_cell_style.alignment = Alignment(wrapText=True, vertical='top', horizontal='left', shrinkToFit=True)
        wb.add_named_style(merged_cell_style)

    file_names = ws['A']
    unique_file_names = list(set([cell.value for cell in file_names if cell.value is not None]))

    for file_name in unique_file_names:
        start_row = None
        end_row = None
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
            if row[0].value == file_name:
                if start_row is None:
                    start_row = row[0].row
                end_row = row[0].row

        if start_row is not None and end_row is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            first_file_name_cell = ws.cell(row=start_row, column=1)
            first_description_cell = ws.cell(row=start_row, column=2)
            first_file_name_cell.style = merged_cell_style_name
            first_description_cell.style = merged_cell_style_name

    set_column_width(ws)

    wb.save(excel_file)

# Основная часть кода
if __name__ == "__main__":
    save_path = str(Path(__file__).parent.resolve())
    folder_path = save_path  # Путь к папке с PDF-файлами
    excel_file = "inductors.xlsx"
    excel_sheet = "Inductors"

    there_is_pdf(folder_path, excel_file, excel_sheet)

    for s in ser1:
        pdf_path = os.path.join(save_path, s)

        # Создание и запись Excel с данными из PDF
        output_excel_path = create_excel_with_file_description(save_path, pdf_path)
        print(f"Processed: {pdf_path}")

        # Обновление и форматирование Excel файла
        update_excel_with_file_paths(output_excel_path, "Sheet1")
