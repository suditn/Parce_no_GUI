import tabula
import pandas as pd
import os
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter
ser1 = []
ser2 = []
def update_excel_with_file_paths(folder_path, excel_file, excel_sheet):
    # Загружаем существующий файл Excel
    wb = load_workbook(excel_file)
    ws = wb[excel_sheet]


    # Создаем словарь для хранения путей к файлам по сериям
    series_paths = {}

    # Рекурсивно сканируем папки и обновляем данные в Excel
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith((".pdf")):
                file_path = os.path.join(root, file)
                # Вычисляем относительный путь
                relative_path = os.path.relpath(file_path, folder_path)
                ser1.append(relative_path)
                # Получаем название серии из пути к файлу
                series_name = os.path.dirname(relative_path)
                series_name = series_name.split(os.sep)[-1]  # Удаляем "Datasheet\" из имени серии
                ser2.append(series_name)
                # Добавляем путь к файлу в словарь для соответствующей серии
                if series_name not in series_paths:
                    series_paths[series_name] = []
                if relative_path not in series_paths[series_name]:
                    series_paths[series_name].append(relative_path)


update_excel_with_file_paths(str(Path(__file__).parent.resolve()), "inductors.xlsx", "Inductors")

for s in ser1:
    print(s)