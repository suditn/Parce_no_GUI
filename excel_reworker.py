import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter
from pathlib import Path

def set_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

def update_excel_with_file_paths(folder_path, excel_file, excel_sheet):
    # Загружаем существующий файл Excel
    wb = load_workbook(excel_file)
    ws = wb[excel_sheet]

    # Создаем словари для хранения путей к файлам по сериям
    series_pdf_paths = {}
    series_zip_paths = {}

    # Рекурсивно сканируем папки и обновляем данные в Excel
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            # Вычисляем относительный путь
            relative_path = os.path.relpath(file_path, folder_path)
            # Получаем название серии из пути к файлу
            series_name = os.path.dirname(relative_path)
            series_name = series_name.split(os.sep)[-1]

            # Добавляем путь к файлу в соответствующий словарь для каждой серии
            if file.endswith(".pdf"):
                if series_name not in series_pdf_paths:
                    series_pdf_paths[series_name] = []
                if relative_path not in series_pdf_paths[series_name]:
                    series_pdf_paths[series_name].append(relative_path)
            elif file.endswith(".zip"):
                if series_name not in series_zip_paths:
                    series_zip_paths[series_name] = []
                if relative_path not in series_zip_paths[series_name]:
                    series_zip_paths[series_name].append(relative_path)

    # Обновляем столбцы "Путь к PDF" и "Путь к ZIP" в файле Excel
    if "Path to the PDF" not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=13, value="Path to the PDF")
    if "Path to the ZIP" not in [cell.value for cell in ws[1]]:
        ws.cell(row=1, column=14, value="Path to the ZIP")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1):
        series_name = row[0].value
        pdf_paths = series_pdf_paths.get(series_name, [])
        zip_paths = series_zip_paths.get(series_name, ["None"])

        # Объединяем пути к файлам в одну строку для каждой серии
        ws.cell(row=row[0].row, column=13, value="; ".join(pdf_paths))
        ws.cell(row=row[0].row, column=14, value="; ".join(zip_paths))

    # Устанавливаем ширину столбцов на основе содержимого
    set_column_width(ws)

    # Сохраняем обновленный файл Excel
    wb.save(excel_file)

def update_csv_with_file_paths(folder_path, csv_file):
    # Загружаем существующий файл CSV
    df = pd.read_csv(csv_file)

    # Создаем словари для хранения путей к файлам по сериям
    series_pdf_paths = {}
    series_zip_paths = {}

    # Рекурсивно сканируем папки и обновляем данные в CSV
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            # Вычисляем относительный путь
            relative_path = os.path.relpath(file_path, folder_path)
            # Получаем название серии из пути к файлу
            series_name = os.path.dirname(relative_path)
            series_name = series_name.split(os.sep)[-1]

            # Добавляем путь к файлу в соответствующий словарь для каждой серии
            if file.endswith(".pdf"):
                if series_name not in series_pdf_paths:
                    series_pdf_paths[series_name] = []
                if relative_path not in series_pdf_paths[series_name]:
                    series_pdf_paths[series_name].append(relative_path)
            elif file.endswith(".zip"):
                if series_name not in series_zip_paths:
                    series_zip_paths[series_name] = []
                if relative_path not in series_zip_paths[series_name]:
                    series_zip_paths[series_name].append(relative_path)

    # Обновляем столбцы "Путь к PDF" и "Путь к ZIP" в файле CSV
    if "Path to the PDF" not in df.columns:
        df["Path to the PDF"] = ""
    if "Path to the ZIP" not in df.columns:
        df["Path to the ZIP"] = ""

    for index, row in df.iterrows():
        series_name = row['Series']
        pdf_paths = series_pdf_paths.get(series_name, [])
        zip_paths = series_zip_paths.get(series_name, ["None"])

        # Объединяем пути к файлам в одну строку для каждой серии
        df.at[index, "Path to the PDF"] = "; ".join(pdf_paths)
        df.at[index, "Path to the ZIP"] = "; ".join(zip_paths)

    # Сохраняем обновленный файл CSV
    df.to_csv("inductors-table.csv", index=False, sep=';')

# Пример использования функций
folder_path = str(Path(__file__).parent.resolve())
update_excel_with_file_paths(folder_path, "inductors.xlsx", "Inductors")
update_csv_with_file_paths(folder_path, "inductors.csv")
