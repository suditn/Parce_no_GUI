import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter

def set_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

def update_excel_with_file_paths(folder_path, excel_file, excel_sheet):
    # Загружаем существующий файл Excel
    wb = load_workbook(excel_file)
    ws = wb[excel_sheet]


    # Создаем словарь для хранения путей к файлам по сериям
    series_paths = {}

    # Рекурсивно сканируем папки и обновляем данные в Excel
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith((".pdf", ".zip")):
                file_path = os.path.join(root, file)
                # Вычисляем относительный путь
                relative_path = os.path.relpath(file_path, folder_path)
                # Получаем название серии из пути к файлу
                series_name = os.path.dirname(relative_path)
                series_name = series_name.split(os.sep)[-1]  # Удаляем "Datasheet\" из имени серии
                # Добавляем путь к файлу в словарь для соответствующей серии
                if series_name not in series_paths:
                    series_paths[series_name] = []
                if relative_path not in series_paths[series_name]:
                    series_paths[series_name].append(relative_path)

    # Обновляем столбец "Путь к файлу" в файле Excel
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        series_name = row[0].value  # Получаем название серии
        if series_name in series_paths:
            cell = ws.cell(row=row[0].row, column=13)  # Столбец "Путь к файлу"
            # Объединяем пути к файлам в одну строку и записываем в ячейку
            cell.value = "\n".join(series_paths[series_name])

    # Объединяем ячейки в столбце "Series" и "Путь к файлу" для каждой серии
    for series_name, file_paths in series_paths.items():
        # Находим начальную и конечную ячейки для текущей серии в столбце "Series"
        start_row = None
        end_row = None
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
            if row[0].value == series_name:
                if start_row is None:
                    start_row = row[0].row
                end_row = row[0].row

        # Объединяем ячейки в столбце "Series" для текущей серии
        if start_row is not None and end_row is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        # Объединяем ячейки в столбце "Путь к файлу" для текущей серии
        if start_row is not None and end_row is not None:
            ws.cell(1, 13).value = "The path to the file"
            ws.merge_cells(start_row=start_row, start_column=13, end_row=end_row, end_column=13)
            ws.merge_cells(start_row=start_row, start_column=12, end_row=end_row, end_column=12)

    # Создаем объект шрифта с помощью свойств из первой ячейки столбца "Series"
    first_cell_font = Font(
        name=ws.cell(row=2, column=1).font.name,  # Имя шрифта
        size=ws.cell(row=2, column=1).font.sz,  # Размер шрифта
        bold=ws.cell(row=2, column=1).font.b,  # Жирный стиль
        italic=ws.cell(row=2, column=1).font.i,  # Курсив
        color=ws.cell(row=2, column=1).font.color,  # Цвет шрифта
        underline=ws.cell(row=2, column=1).font.underline,  # Подчеркивание
        strike=ws.cell(row=2, column=1).font.strike,  # Зачеркивание
    )

    # Создаем именованный стиль с уменьшенной высотой строки
    merged_cell_style_name = "merged_cell_style"
    if merged_cell_style_name not in wb.named_styles:
        merged_cell_style = NamedStyle(name=merged_cell_style_name, font=first_cell_font)
        merged_cell_style.alignment = Alignment(wrapText=True, vertical='top', horizontal='left', shrinkToFit=True)
        wb.add_named_style(merged_cell_style)

    # Применяем стиль к объединенным ячейкам
    for merged_cell_range in ws.merged_cells.ranges:
        for row in ws.iter_rows(
            min_row=merged_cell_range.min_row,
            max_row=merged_cell_range.max_row,
            min_col=merged_cell_range.min_col,
            max_col=merged_cell_range.max_col
        ):
            for cell in row:
                cell.style = merged_cell_style_name

    # Устанавливаем ширину столбцов на основе содержимого
    set_column_width(ws)

    # Сохраняем обновленный файл Excel
    wb.save(excel_file)

# Пример использования функции
update_excel_with_file_paths("C:\\Users\\stud\\Desktop\\Parceprob2-master\\", "inductors.xlsx", "Inductors")