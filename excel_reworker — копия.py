import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter
import pandas as pd
import logging
import csv

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

save_path = str(Path(__file__).parent.resolve())
logging.info(f'Текущая директория: {save_path}')

def set_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

def is_number(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def convert_to_number(value):
    try:
        return float(value)
    except ValueError:
        return value

def update_excel_with_file_paths(folder_path, excel_file, excel_sheet):
    logging.info(f'Обработка файла Excel: {excel_file}')
    wb = load_workbook(excel_file)
    ws = wb[excel_sheet]

    last_column = ws.max_column

    series_paths = {}

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith((".pdf", ".zip")):
                file_path = os.path.join(root, file)
                relative_path = os.path.relpath(file_path, folder_path)
                series_name = os.path.dirname(relative_path).split(os.sep)[-1]
                if series_name not in series_paths:
                    series_paths[series_name] = []
                if relative_path not in series_paths[series_name]:
                    series_paths[series_name].append(relative_path)

    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        series_name = row[0].value
        if series_name in series_paths:
            cell = ws.cell(row=row[0].row, column=last_column + 1)
            cell.value = "; ".join(series_paths[series_name])

    for series_name, file_paths in series_paths.items():
        start_row = None
        end_row = None
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
            if row[0].value == series_name:
                if start_row is None:
                    start_row = row[0].row
                end_row = row[0].row

        if start_row is not None and end_row is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        if start_row is not None and end_row is not None:
            ws.cell(1, last_column + 1).value = "The path to the file"
            ws.merge_cells(start_row=start_row, start_column=last_column + 1, end_row=end_row, end_column=last_column + 1)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str) and is_number(cell.value):
                    cell.value = convert_to_number(cell.value)

    first_cell_font = Font(
        name=ws.cell(row=2, column=1).font.name,
        size=ws.cell(row=2, column=1).font.sz,
        bold=ws.cell(row=2, column=1).font.b,
        italic=ws.cell(row=2, column=1).font.i,
        color=ws.cell(row=2, column=1).font.color,
        underline=ws.cell(row=2, column=1).font.underline,
        strike=ws.cell(row=2, column=1).font.strike,
    )

    merged_cell_style_name = "merged_cell_style"
    if merged_cell_style_name not in wb.named_styles:
        merged_cell_style = NamedStyle(name=merged_cell_style_name, font=first_cell_font)
        merged_cell_style.alignment = Alignment(wrapText=True, vertical='top', horizontal='left', shrinkToFit=True)
        wb.add_named_style(merged_cell_style)

    for merged_cell_range in ws.merged_cells.ranges:
        for row in ws.iter_rows(
            min_row=merged_cell_range.min_row,
            max_row=merged_cell_range.max_row,
            min_col=merged_cell_range.min_col,
            max_col=merged_cell_range.max_col
        ):
            for cell in row:
                cell.style = merged_cell_style_name

    set_column_width(ws)

    wb.save(excel_file)
    logging.info(f'Обновленный файл Excel сохранен: {excel_file}')

def update_csv_with_file_paths(folder_path, csv_file):
    logging.info(f'Обработка файла CSV: {csv_file}')
    df = pd.read_csv(csv_file)

    last_column = len(df.columns)

    series_paths = {}

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith((".pdf", ".zip")):
                file_path = os.path.join(root, file)
                relative_path = os.path.relpath(file_path, folder_path)
                series_name = os.path.dirname(relative_path).split(os.sep)[-1]
                if series_name not in series_paths:
                    series_paths[series_name] = []
                if relative_path not in series_paths[series_name]:
                    series_paths[series_name].append(relative_path)

    df['The path to the file'] = df.iloc[:, 0].map(series_paths).apply(lambda x: "; ".join(x) if isinstance(x, list) else '')

    for column in df.select_dtypes(include=[object]).columns:
        df[column] = df[column].apply(lambda x: convert_to_number(x) if is_number(x) else x)

    df.to_csv(csv_file, index=False, quoting=csv.QUOTE_NONE, escapechar=' ', sep=';')
    logging.info(f'Обновленный файл CSV сохранен: {csv_file}')

def process_files_in_directory(directory):
    logging.info(f'Сканирование директории: {directory}')
    for file_name in os.listdir(directory):
        file_path = os.path.join(directory, file_name)
        if file_name.endswith(".xlsx"):
            logging.info(f'Найден файл Excel: {file_name}')
            excel_sheet_name = file_name.split('.')[-2]
            update_excel_with_file_paths(directory, file_path, excel_sheet_name)
        elif file_name.endswith(".csv"):
            logging.info(f'Найден файл CSV: {file_name}')
            update_csv_with_file_paths(directory, file_path)

process_files_in_directory(save_path)
